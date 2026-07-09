"""
Trim Decision Pricing Model (scalable-model)
============================================
Reads a Comact TrimExpert AllProducts.xml export and builds, per species + thickness,
a single value ladder of the grades the Comact runs, plus the trim decision matrix.

Everything is done through the PRICES, because the price grid is the only thing the Comact
consumes. There is no decision-side rule: the trim matrix is pure value, exactly what the
Comact does with these prices. So the two protections below are built into the numbers:

- Even-length protection is a consequence of GAP. As long as each one-grade price step stays
  in the safe band, no even board out-values a shorter odd board one grade up, so a 14 or 16
  never trims down to a 13 or 15. See GAP below.
- Freezing the long lengths (13+) is done by flattening their prices: FREEZE_FROM sets every
  grade at those lengths to the top price, so length wins and grade gives no reason to trim.

Each species is one ordered ladder of tiers, highest value first. A tier holds one or more
grades; grades in the same tier are tied (same price), so the Comact never trims between them.
Global rule: Select is tied into the 1 Common tier, so 1 Common never trims up to Select.

Adjustability: tiers are data (order + ties + which lengths carry variance). Tune trims with
GAP, ODD_DISCOUNT, SIX_DISCOUNT, and FREEZE_FROM. Add back an excluded grade by putting its
name in a tier and un-commenting one line in that species' parser.

Batch mode
----------
    py trim_model.py AllProducts.xml --all --max-length 12 --thickness 4/4 6/4

Writes one .xlsx (default trim_models.xlsx) with a formatted visualizer sheet per
species x thickness, matching the floor visualizer layout: price/1000 grid, $/board
grid with SM row, and the three decision blocks with red conditional highlighting
(any positive gain = the trim fires). Single-combo CSV mode still works:
    py trim_model.py AllProducts.xml TUL 4/4
"""
import xml.etree.ElementTree as ET
import csv
from collections import defaultdict

# ============================ POLICY CONFIG ============================
ODD_DISCOUNT = 0.06   # even/odd lever for variance tiers. MUST stay < 1/15 = 0.0667.
SIX_DISCOUNT = 0.09   # 6' low-point discount on variance tiers
MARGIN       = 0.0    # pure value: any positive gain upgrades. 0 matches the Comact.

# GAP is the one-grade price step: each tier is priced at GAP x the tier above it.
# It is the even-length protection. Keep it in the band below:
#   * GAP >= 0.882  so a 16 (or 14) never out-trims to a shorter odd one grade up.
#   * GAP <= 0.912  so the odd upgrades at 9/11/13/15 still fire.
# 0.89 sits in the middle (about a 12.4% one-grade gap).
GAP = 0.89

# FREEZE_FROM: set to a length (e.g. 13) to flatten every grade at that length and longer to
# the top price. Those lengths then never trim (length wins, grade is equalized). None = off.
FREEZE_FROM = None

LENGTHS = list(range(6, 17))

# Visualizer grid: fixed number of grade slots per sheet. Unused slots render as the
# yellow "0" rows, and they are what generates the always-red subgrade-escape rows in
# the 1 Grade / 2 Grades blocks (value of the tier above minus zero).
GRADE_SLOTS = 9

# Each ladder: list of tiers, HIGHEST value first.  tier = (has_even_odd_variance, [grade names]).
# Same tier = tied price. Prices are generated from GAP, so only the ORDER and the TIES live here.

# ---- Hard maple: no 1W / 2W / FAS Brown (excluded). Sap and unselected on one ladder. ----
HMW = [
    (True,  ['FAS S']),
    (True,  ['1 Common']),                 # 1 Common unselected
    (True,  ['Sap Select', '1 Common Sap']),
    (True,  ['2 Common Sap']),
    (True,  ['3A Sap']),                   # no raw grade in export yet; kept for later
    (True, ['1 Common Brown', '2 Common', '3A Common', '3B Common', 'Subgrade']),
]

# ---- Soft maple: hard maple plus plain Select (tied to 1 Common) and Wormy (between
#      2 Common Sap and 3A Sap). ----
SMA = [
    (True,  ['FAS S']),
    (True,  ['1 Common', 'Select']),
    (True,  ['Sap Select', '1 Common Sap']),
    (True,  ['2 Common Sap']),
    (True,  ['Wormy']),
    (True,  ['3A Sap']),                   # no raw grade in export yet; kept for later
    (True, ['1 Common Brown', '2 Common', '3A Common', '3B Common', 'Subgrade']),
]

# ---- Ash: no color sort. All color grades collapse to their base grade. ----
ASH = [
    (True,  ['FAS']),
    (True,  ['1 Common', 'Select']),
    (True,  ['2 Common']),
    (True,  ['3A Common']),
    (True, ['3B Common']),
    (True, ['Subgrade']),                 # SG
    (True, ['Pallet']),                   # no raw grade in export yet; kept for later
]

# ---- Cherry: heartwood on top. Sap is not wanted, so it prices into the 1 Common tier. ----
CHERRY = [
    (True,  ['FAS 90-50', 'FAS Red']),
    (True,  ['Select 90-50', '1 Common 90-50', '1 Common', 'Select', 'FAS Sap']),
    (True,  ['2 Common']),
    (True,  ['3A Common']),
    (True, ['Subgrade']),                 # SG
    (True, ['3B Common', 'Pallet']),
]

# ---- Red oak: no color sort, no quarter sawn, no fas 10in (excluded). ----
ROK = [
    (True,  ['FAS', 'FAS Stain']),
    (True,  ['1 Common', 'Select']),
    (True,  ['1&2 Common']),
    (True,  ['2 Common']),
    (True,  ['3A Common']),
    (True, ['3B Common', 'Subgrade']),
]

# ---- White oak: like red oak, plus Character between 2 Common and 3A. ----
WOK = [
    (True,  ['FAS', 'FAS Stain']),
    (True,  ['1 Common', 'Select']),
    (True,  ['2 Common']),
    (True,  ['Character']),
    (True,  ['3A Common']),
    (True, ['3B Common', 'Subgrade']),
]

# ---- Walnut: keep everything, do not want to trim. All tiers flat (no even-length trims).
#      No subgrade. ----
WALNUT = [
    # One tied price for every grade: nothing to gain by trimming, so walnut never trims.
    (True, ['FAS', '1 Common', 'Select', '2 Common', '3A Common', '3B Common']),
]

# ---- Basswood / birch / tulip: simple, no subgrade. ----
PLAIN = [
    (True,  ['FAS']),
    (True,  ['1 Common', 'Select']),
    (True,  ['2 Common']),
    (True,  ['3A Common']),
    (True, ['3B Common']),
]


# ---- Parsers: raw grade string -> ladder grade name, or None to exclude. ----
def canon_hmw(g):
    g = g.upper()
    if g.startswith('VENEER') or 'WORMY' in g: return None
    if g.startswith('FAS1W') or g.startswith('FAS2W') or g.startswith('FASB'): return None  # excluded: re-add to a tier to restore
    if g.startswith('FASS'): return 'FAS S'
    if g.startswith('FAS'):  return None
    if '1COM' in g and 'SAP' in g: return '1 Common Sap'
    if '2COM' in g and 'SAP' in g: return '2 Common Sap'
    if g.startswith('SEL') and 'SAP' in g: return 'Sap Select'
    if g.startswith('SEL'): return None
    if g.startswith('1COMB'): return '1 Common Brown'
    if g.startswith('1COM'):  return '1 Common'
    if g.startswith('2COM'):  return '2 Common'
    if g.startswith('3ACOM'): return '3A Common'
    if g.startswith('3B'):    return '3B Common'
    if g.startswith('SUBG'):  return 'Subgrade'
    return None

def canon_sma(g):
    g = g.upper()
    if g.startswith('VENEER'): return None
    if 'WORMY' in g: return 'Wormy'
    if g.startswith('FAS1W') or g.startswith('FAS2W') or g.startswith('FASB'): return None  # excluded
    if g.startswith('FASS'): return 'FAS S'
    if g.startswith('FAS'):  return None
    if '1COM' in g and 'SAP' in g: return '1 Common Sap'
    if '2COM' in g and 'SAP' in g: return '2 Common Sap'
    if g.startswith('SEL') and 'SAP' in g: return 'Sap Select'
    if g.startswith('SEL'): return 'Select'
    if g.startswith('1COMB'): return '1 Common Brown'
    if g.startswith('1COM'):  return '1 Common'
    if g.startswith('2COM'):  return '2 Common'
    if g.startswith('3ACOM'): return '3A Common'
    if g.startswith('3B'):    return '3B Common'
    if g.startswith('SUBG'):  return 'Subgrade'
    return None

def canon_ash(g):
    g = g.upper()
    if g.startswith('VENEER'): return None
    if g.startswith('FAS'): return 'FAS'        # no color sort: FAS1W/FAS2W/FASS/FASB all fold to FAS
    if g.startswith('SEL'): return 'Select'     # SEL/SEL1W/SEL2W/SEL SAP all fold to Select
    if g.startswith('1COM'): return '1 Common'
    if g.startswith('2COM'): return '2 Common'
    if g.startswith('3ACOM'): return '3A Common'
    if g.startswith('3B'): return '3B Common'
    if g.startswith('SUBG'): return 'Subgrade'
    return None

def canon_cherry(g):
    g = g.upper()
    if '2CHPY' in g or g.startswith('VENEER'): return None
    if g.startswith('FASR'):     return 'FAS Red'
    if g.startswith('FAS9050'):  return 'FAS 90-50'
    if g.startswith('FASSAP'):   return 'FAS Sap'
    if g.startswith('FAS'):      return None
    if g.startswith('SEL9050'):  return 'Select 90-50'
    if g.startswith('SEL') and 'SAP' in g: return 'FAS Sap'   # sap select prices with the sap tier
    if g.startswith('SEL'):      return 'Select'
    if g.startswith('1COM9050'): return '1 Common 90-50'
    if g.startswith('1COM'):     return '1 Common'
    if g.startswith('2COM'):     return '2 Common'
    if g.startswith('3ACOM'):    return '3A Common'
    if g.startswith('3B'):       return '3B Common'
    if g.startswith('SUBG'):     return 'Subgrade'
    return None

def canon_rok(g):
    g = g.upper()
    if 'QUARTER' in g: return None              # excluded: no quarter sawn
    if g.startswith('FAS10'): return None       # excluded: no fas 10in
    if '2CHPY' in g or g.startswith('VENEER'): return None
    if 'STAIN' in g: return 'FAS Stain'
    if g.startswith('FAS'): return 'FAS'
    if g.startswith('SEL'): return 'Select'
    if g.startswith('1C/2C'): return '1&2 Common'
    if g.startswith('1COM'): return '1 Common'
    if g.startswith('2COM'): return '2 Common'
    if g.startswith('3ACOM'): return '3A Common'
    if g.startswith('3B'): return '3B Common'
    if g.startswith('SUBG'): return 'Subgrade'
    return None

def canon_wok(g):
    g = g.upper()
    if 'QUARTER' in g: return None              # excluded: no quarter sawn
    if '2CHPY' in g or g.startswith('VENEER'): return None
    if 'STAIN' in g: return 'FAS Stain'
    if g.startswith('CHAR'): return 'Character'
    if g.startswith('FAS'): return 'FAS'
    if g.startswith('SEL'): return 'Select'
    if g.startswith('1C/2C'): return '1 Common'
    if g.startswith('1COM'): return '1 Common'
    if g.startswith('2COM'): return '2 Common'
    if g.startswith('3ACOM'): return '3A Common'
    if g.startswith('3B'): return '3B Common'
    if g.startswith('SUBG'): return 'Subgrade'
    return None

def canon_walnut(g):
    g = g.upper()
    if g.startswith('VENEER'): return None
    if g.startswith('FAS'): return 'FAS'        # incl FAS OPT 6-7' (short FAS prices as FAS)
    if g.startswith('SEL'): return 'Select'
    if g.startswith('1COM'): return '1 Common'
    if g.startswith('2COM'): return '2 Common'
    if g.startswith('3ACOM'): return '3A Common'
    if g.startswith('3B'): return '3B Common'
    if g.startswith('SUBG'): return None        # no subgrade
    return None

def canon_plain(g):
    g = g.upper()
    if g.startswith('VENEER') or 'WORMY' in g or 'CHAR' in g or 'STAIN' in g: return None
    if g.startswith('FAS'): return 'FAS'
    if g.startswith('SEL'): return 'Select'
    if g.startswith('1C/2C'): return '1 Common'
    if g.startswith('1COM'): return '1 Common'
    if g.startswith('2COM'): return '2 Common'
    if g.startswith('3ACOM'): return '3A Common'
    if g.startswith('3B'): return '3B Common'
    if g.startswith('SUBG'): return None        # no subgrade
    return None

SPECIES = {
    'HMW':      (HMW,    canon_hmw),
    'SMA':      (SMA,    canon_sma),
    'ASH':      (ASH,    canon_ash),
    'CHERRY':   (CHERRY, canon_cherry),
    'ROK':      (ROK,    canon_rok),
    'WOK':      (WOK,    canon_wok),
    'WALNUT':   (WALNUT, canon_walnut),
    'BASSWOOD': (PLAIN,  canon_plain),
    'BIRCH':    (PLAIN,  canon_plain),
    'TUL':      (PLAIN,  canon_plain),
}
# =======================================================================

def load_raw(path):
    """Return {(species, thick): set(raw grade strings)}."""
    root = ET.parse(path).getroot()
    combo = defaultdict(set)
    for b in root.findall('Board'):
        d = {c.tag: (c.text or '') for c in b}
        species = d['name'].strip().split()[-1]
        combo[(species, d['thick'])].add(d['grade'])
    return combo

def tiers_for(combo, species, thick):
    """Return the present tiers as [{value, variance, names}], highest first. Prices come
    from GAP: the top present tier is 1.0, each lower present tier is GAP x the one above."""
    spec = SPECIES.get(species)
    if not spec: return []
    ladder, canon = spec
    raw = combo.get((species, thick), set())
    name_to_tier = {}
    for i, (var, names) in enumerate(ladder):
        for n in names: name_to_tier[n] = i
    present = defaultdict(list)
    for g in raw:
        nm = canon(g)
        if nm is None or nm not in name_to_tier: continue
        idx = name_to_tier[nm]
        if nm not in present[idx]: present[idx].append(nm)
    tiers, rank = [], 0
    for i, (var, names) in enumerate(ladder):
        if i in present:
            ordered = [n for n in names if n in present[i]]
            tiers.append({'value': GAP ** rank, 'variance': var, 'names': ordered})
            rank += 1
    return tiers

def price(tier, L, scale=1.0):
    if FREEZE_FROM is not None and L >= FREEZE_FROM:
        return 1.0 * scale                            # frozen: flat top price -> length wins, no trim
    base = tier['value'] * scale
    if not tier['variance']: return base              # flat: same $/MBF every length
    if L == 6: return base * (1 - SIX_DISCOUNT)
    if L % 2 == 0: return base
    return base * (1 - ODD_DISCOUNT)

def board_value(tier, L, scale=1.0, width=12, thick_dec=1.0):
    return price(tier, L, scale) / 1000 * thick_dec * width * L / 12

def decision_blocks(tiers, margin=MARGIN, width=12, thick_dec=1.0):
    """Pure value, exactly what the Comact does with these prices: a lower tier climbs to a
    higher tier when the shorter board at the higher tier is worth more."""
    def bv(t, L): return board_value(t, L, 1.0, width, thick_dec)
    def reds(cur, up):
        return [L for L in range(7, 17) if (bv(up, L-1) - bv(cur, L)) / bv(cur, L) > margin]
    def label(t): return ' / '.join(t['names'])
    res = {'same': {}, 'up1': {}, 'up2': {}}
    for t in tiers:
        res['same'][label(t)] = reds(t, t)
    for i in range(1, len(tiers)):
        res['up1'][f"{label(tiers[i])} -> {label(tiers[i-1])}"] = reds(tiers[i], tiers[i-1])
    for i in range(2, len(tiers)):
        res['up2'][f"{label(tiers[i])} -> {label(tiers[i-2])}"] = reds(tiers[i], tiers[i-2])
    return res

def write_report_csv(combo, species, thick, path, scale=600.0, width=12):
    tiers = tiers_for(combo, species, thick)
    labels = [' / '.join(t['names']) for t in tiers]
    rows = [[f'{species} {thick} pricing model  (relative, illustrative scale top = {scale:g})'],
            ['tiers'] + labels, []]
    rows.append(['PRICE / 1000'] + LENGTHS)
    for t in tiers:
        rows.append([' / '.join(t['names'])] + [round(price(t, L, scale)) for L in LENGTHS])
    rows += [[], [f'$ / BOARD (width {width})'] + LENGTHS]
    for t in tiers:
        rows.append([' / '.join(t['names'])] + [round(board_value(t, L, scale, width), 2) for L in LENGTHS])
    res = decision_blocks(tiers, width=width)
    for title, blk in (('SAME TIER', 'same'), ('1 TIER UP', 'up1'), ('2 TIERS UP', 'up2')):
        rows += [[], [title + '  (R = trim fires)'] + list(range(7, 17))]
        for k, r in res[blk].items():
            rows.append([k] + ['R' if L in r else '.' for L in range(7, 17)])
    with open(path, 'w', newline='') as f:
        csv.writer(f).writerows(rows)
    return labels

# ============================ XLSX VISUALIZER (batch mode) ============================
# One formatted sheet per species x thickness, matching the floor visualizer:
#   A1  title, price/1000 grid, black divider, $/board grid with SM row and the yellow
#   zero slots, then the three decision blocks to the right with red highlighting on any
#   positive gain (red = the trim fires), and the Ideal State Map box.
# All prices are ROUNDED to whole $/MBF first and every $/board and gain figure is
# computed from the rounded price, so the workbook matches what the grid displays.

def thick_decimal(thick):
    try:
        num, den = thick.split('/')
        return float(num) / float(den)
    except (ValueError, ZeroDivisionError):
        return 1.0

def write_workbook(combo, out_path, species_list, thicknesses, max_length=16,
                   width=8, scale=600.0, apply_thickness=True):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    YELLOW = PatternFill('solid', start_color='FFFF00')
    # For conditional-format (dxf) fills Excel uses the BACKGROUND color of a
    # solid fill, so both start_color and end_color must be set or real Excel
    # renders no highlight (previews/LibreOffice show fgColor and look fine).
    RED     = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    RED_FONT = Font(color='9C0006')                       # dark red, matches Excel "Bad"
    THIN   = Side(style='thin', color='000000')
    THICK  = Side(style='medium', color='000000')
    BOX    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal='center')
    FMT_PRICE = '0'
    FMT_BOARD = '0.00" $"'
    FMT_SAME  = '"$"0.000;-"$"0.000'
    FMT_GAIN  = '"$"0.00;-"$"0.00'

    # ---- fonts: change these to restyle the whole workbook ----
    FONT_NAME = 'Aptos Narrow'
    BODY  = Font(name=FONT_NAME, size=11)                 # everything not listed below
    BOLD  = Font(name=FONT_NAME, size=11, bold=True)      # yellow grade/length rows
    HDR   = Font(name=FONT_NAME, size=14)                 # section headers
    HDR_B = Font(name=FONT_NAME, size=14, bold=True)
    TITLE = Font(name=FONT_NAME, size=20, bold=True)      # 'SMA 4/4' title

    def rich(ws, row, col, bold_part, rest):
        cell = ws.cell(row=row, column=col)
        cell.font = HDR
        try:
            from openpyxl.cell.rich_text import CellRichText, TextBlock
            from openpyxl.cell.text import InlineFont
            cell.value = CellRichText(
                TextBlock(InlineFont(rFont=FONT_NAME, sz=14, b=True), bold_part), rest)
        except Exception:
            cell.value = bold_part + rest
            cell.font = HDR_B
        return cell

    lengths = list(range(6, max_length + 1))
    trims   = list(range(7, max_length + 1))      # trim-from lengths (right blocks)
    n_len   = len(lengths)
    c_label = 1                                    # column A
    c_len0  = 2                                    # first length column (B)
    c_r0    = c_len0 + n_len + 1                   # right block start (gap column between)

    # fixed row skeleton (matches the visualizer)
    R_TITLE, R_PLBL, R_PHDR, R_P0 = 1, 4, 5, 6     # price table starts at row 6
    R_SEP  = 14                                    # black divider band
    R_WID, R_LEN, R_SM, R_G0 = 16, 17, 18, 19      # $/board table, 9 grade slots 19..27
    R_SG_H, R_SG0 = 2, 3                           # same grade: header + 9 rows (3..11)
    R_U1_H, R_U10 = 13, 14                         # 1 grade:    header + 8 rows (14..21)
    R_U2_H, R_U20 = 22, 23                         # 2 grades:   header + 7 rows (23..29)
    R_MAP_H, R_MAP0 = 30, 31                       # ideal state map: header + 3 rows

    wb = Workbook()
    try:
        wb._named_styles['Normal'].font = BODY   # workbook-wide default (body) font
    except Exception:
        pass
    wb.remove(wb.active)
    written = []

    for species in species_list:
        for thick in thicknesses:
            tiers = tiers_for(combo, species, thick)
            if not tiers:
                continue
            tdec = thick_decimal(thick) if apply_thickness else 1.0
            ws = wb.create_sheet(f"{species} {thick.replace('/', '-')}")
            ws.column_dimensions['A'].width = 30
            for i in range(n_len):
                ws.column_dimensions[get_column_letter(c_len0 + i)].width = 8.5
            ws.column_dimensions[get_column_letter(c_len0 + n_len)].width = 3
            for i in range(len(trims)):
                ws.column_dimensions[get_column_letter(c_r0 + i)].width = 10

            # rounded $/MBF per tier slot, then $/board from the rounded price
            p = [{L: round(price(t, L, scale)) for L in lengths} for t in tiers]
            def bv(slot, L):
                if slot < 0 or slot >= len(tiers): return 0.0
                return p[slot][L] / 1000.0 * tdec * width * L / 12.0
            labels = [' / '.join(t['names']) for t in tiers]

            # ---- title ----
            c_mid = c_len0 + max(n_len // 2 - 1, 0)   # ~centered over the table (D for 6-12)
            tc = ws.cell(row=R_TITLE, column=c_mid, value=f'{species} {thick}')
            tc.font = TITLE; tc.alignment = CENTER

            # ---- price/1000 table ----
            lc = ws.cell(row=R_PLBL, column=c_mid, value='price/1000')
            lc.alignment = CENTER
            h = ws.cell(row=R_PHDR, column=c_label, value='Grade')
            h.fill = YELLOW; h.font = BOLD; h.border = BOX
            for i, L in enumerate(lengths):
                c = ws.cell(row=R_PHDR, column=c_len0 + i, value=L)
                c.fill = YELLOW; c.font = BOLD; c.border = BOX; c.alignment = CENTER
            for r, t in enumerate(tiers):
                lab = ws.cell(row=R_P0 + r, column=c_label, value=labels[r])
                lab.fill = YELLOW; lab.border = BOX
                lab.alignment = Alignment(wrap_text=True, vertical='center')
                if len(labels[r]) > 34: ws.row_dimensions[R_P0 + r].height = 27
                for i, L in enumerate(lengths):
                    c = ws.cell(row=R_P0 + r, column=c_len0 + i, value=p[r][L])
                    c.number_format = FMT_PRICE; c.border = BOX

            # ---- divider: bold border along the top of the $/board table ----
            for col in range(c_label, c_len0 + n_len):
                ws.cell(row=R_SEP + 1, column=col).border = Border(top=THICK)

            # ---- $/board table ----
            ws.cell(row=R_WID, column=c_len0 + 2, value='width').alignment = CENTER
            ws.cell(row=R_WID, column=c_len0 + 3, value=f"{width:g}.''").alignment = CENTER
            h = ws.cell(row=R_LEN, column=c_label, value='length')
            h.fill = YELLOW; h.font = BOLD; h.border = BOX
            for i, L in enumerate(lengths):
                c = ws.cell(row=R_LEN, column=c_len0 + i, value=L)
                c.fill = YELLOW; c.font = BOLD; c.border = BOX; c.alignment = CENTER
            s = ws.cell(row=R_SM, column=c_label, value='SM')
            s.fill = YELLOW; s.border = BOX
            for i, L in enumerate(lengths):
                c = ws.cell(row=R_SM, column=c_len0 + i, value=round(width * L / 12))
                c.border = BOX; c.alignment = CENTER
            for slot in range(GRADE_SLOTS):
                row = R_G0 + slot
                lab = ws.cell(row=row, column=c_label,
                              value=labels[slot] if slot < len(tiers) else 0)
                lab.fill = YELLOW; lab.border = BOX
                lab.alignment = Alignment(wrap_text=True, vertical='center')
                if slot < len(tiers) and len(labels[slot]) > 34:
                    ws.row_dimensions[row].height = 27
                for i, L in enumerate(lengths):
                    # full precision, displayed as 2 decimals: the decision
                    # formulas depend on the unrounded values (goal shows -$0.448)
                    c = ws.cell(row=row, column=c_len0 + i, value=bv(slot, L))
                    c.number_format = FMT_BOARD; c.border = BOX

            # ---- decision blocks ----
            # Decision cells are live formulas against the $/board grid, like the
            # original visualizer: keep at L (col of L-1, same grade row) minus
            # trim to L-1 at a grade `up` slots better (row R_G0+j+up, col of L).
            def block(hdr_row, row0, n_rows, bold_part, rest, fmt, up):
                rich(ws, hdr_row, c_r0, bold_part, rest)
                for j in range(n_rows):
                    for i, L in enumerate(trims):
                        keep = f'{get_column_letter(c_len0 + L - 1 - lengths[0])}{R_G0 + j}'
                        trim = f'{get_column_letter(c_len0 + L - lengths[0])}{R_G0 + j + up}'
                        c = ws.cell(row=row0 + j, column=c_r0 + i,
                                    value=f'={keep}-{trim}')
                        c.number_format = fmt; c.border = BOX
                first = f'{get_column_letter(c_r0)}{row0}'
                last  = f'{get_column_letter(c_r0 + len(trims) - 1)}{row0 + n_rows - 1}'
                ws.conditional_formatting.add(
                    f'{first}:{last}',
                    CellIsRule(operator='greaterThan', formula=['0'],
                               fill=RED, font=RED_FONT))

            block(R_SG_H, R_SG0, GRADE_SLOTS,
                  'Price in same Grade', ' (red will trim in same grade category)',
                  FMT_SAME, 0)
            block(R_U1_H, R_U10, GRADE_SLOTS - 1,
                  '1 Grade', " - Red will trim 1' to upgrade a single grade",
                  FMT_GAIN, 1)
            block(R_U2_H, R_U20, GRADE_SLOTS - 2,
                  '2 Grades', " - Red will trim 1' to recover 2 grades",
                  FMT_GAIN, 2)

            # ---- ideal state map ----
            mh = ws.cell(row=R_MAP_H, column=c_r0, value='Ideal State Map')
            mh.font = HDR_B
            odd_up = [L for L in trims if L % 2 == 1 and L >= 9]   # 9/11/13/15 odd upgrades
            for j in range(2):
                for i, L in enumerate(trims):
                    if L in odd_up:
                        c = ws.cell(row=R_MAP0 + j, column=c_r0 + i, value=0.01)
                        c.number_format = FMT_GAIN; c.fill = RED
            for i, L in enumerate(trims):
                c = ws.cell(row=R_MAP0 + 2, column=c_r0 + i, value=0.01)
                c.number_format = FMT_GAIN; c.fill = RED
            for row in range(R_MAP_H, R_MAP0 + 3):
                for col in range(c_r0, c_r0 + len(trims)):
                    cell = ws.cell(row=row, column=col)
                    b = {'left': THIN, 'right': THIN, 'top': THIN, 'bottom': THIN}
                    if row == R_MAP_H: b['top'] = THICK
                    if row == R_MAP0 + 2: b['bottom'] = THICK
                    if col == c_r0: b['left'] = THICK
                    if col == c_r0 + len(trims) - 1: b['right'] = THICK
                    cell.border = Border(**{k: v for k, v in b.items()})

            # ---- font normalization ----
            # Cells with any direct format (border, number format, fill) get their
            # own style record pinned to font 0 (Calibri); real Excel uses that
            # instead of the Normal-style default. Stamp the body font on every
            # written cell that wasn't given an explicit Aptos font above.
            for xrow in ws.iter_rows():
                for xc in xrow:
                    if xc.font is None or xc.font.name != FONT_NAME:
                        xc.font = BODY

            written.append(f'{species} {thick}')

    if not written:
        return []
    wb.save(out_path)
    return written

if __name__ == '__main__':
    import argparse, sys
    ap = argparse.ArgumentParser(description='Comact trim decision pricing model')
    ap.add_argument('xml', help='AllProducts.xml export')
    ap.add_argument('species', nargs='?', help='single-combo CSV mode: species code')
    ap.add_argument('thick', nargs='?', help='single-combo CSV mode: thickness, e.g. 4/4')
    ap.add_argument('--all', action='store_true',
                    help='batch mode: every species in the book -> one formatted .xlsx')
    ap.add_argument('--thickness', nargs='+', metavar='T',
                    help='thicknesses for batch mode, e.g. --thickness 4/4 6/4 '
                         '(default: every thickness in the export)')
    ap.add_argument('--max-length', type=int, default=16, choices=range(8, 17),
                    help='longest length column, e.g. 12 (default 16)')
    ap.add_argument('--width', type=float, default=8, help='board width in inches (default 8)')
    ap.add_argument('--scale', type=float, default=600.0,
                    help='illustrative top even price (default 600)')
    ap.add_argument('--out', default='trim_models.xlsx', help='batch output workbook')
    ap.add_argument('--flat-thickness', action='store_true',
                    help='do NOT multiply $/board by thickness (6/4 x1.5 etc.)')
    args = ap.parse_args()

    combo = load_raw(args.xml)

    if args.all:
        thicknesses = args.thickness or sorted({t for (_, t) in combo},
                                               key=lambda s: thick_decimal(s))
        written = write_workbook(combo, args.out, list(SPECIES), thicknesses,
                                 max_length=args.max_length, width=args.width,
                                 scale=args.scale,
                                 apply_thickness=not args.flat_thickness)
        if not written:
            print('No known species/thickness combos found in export.'); sys.exit(1)
        skipped = [f'{s} {t}' for s in SPECIES for t in thicknesses
                   if f'{s} {t}' not in written]
        print(f'wrote {args.out}: {len(written)} sheets')
        print('  ' + ' | '.join(written))
        if skipped:
            print('skipped (not in export): ' + ' | '.join(skipped))
        sys.exit(0)

    # ---- legacy single-combo CSV mode ----
    if args.species and args.thick:
        species, thick = args.species, args.thick
    else:
        species, thick = max((k for k in combo if k[0] in SPECIES),
                             key=lambda k: len(combo[k]), default=(None, None))
    if species is None:
        print('No known species found in export.'); sys.exit(0)
    out = f'sample_{species}_{thick.replace("/", "-")}.csv'
    labels = write_report_csv(combo, species, thick, out)
    print(f'{species} {thick} tiers: ' + ' | '.join(labels))
    print(f'wrote {out}')
